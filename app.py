"""
app.py — Generador PESOS con procesamiento asíncrono
Flujo:
  1. POST /iniciar   → guarda el fichero, lanza hilo, devuelve job_id (rápido)
  2. GET  /estado/<id> → devuelve progreso o "listo"
  3. GET  /descargar/<id> → devuelve el xlsx generado
"""
from flask import Flask, request, send_file, render_template_string, jsonify
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
from collections import defaultdict
import io, os, gc, uuid, threading, time, traceback
from datetime import datetime

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

# ── Store de trabajos en memoria ──────────────────────────────────────────────
# {job_id: {status, progress, message, result_buf, fname, error, ts}}
JOBS = {}
JOBS_LOCK = threading.Lock()

def cleanup_old_jobs():
    """Elimina trabajos con más de 1 hora."""
    now = time.time()
    with JOBS_LOCK:
        old = [jid for jid, j in JOBS.items() if now - j['ts'] > 3600]
        for jid in old:
            del JOBS[jid]

# ── Estilos ───────────────────────────────────────────────────────────────────
BLK='000000';WHT='FFFFFF';ORG='E97132';GRY='F2F2F2';GRN='C6EFCE';RED='FFC7CE'
FMT_EUR='_-* #,##0.00\\ "€"_-;\\-* #,##0.00\\ "€"_-;_-* "-"??\\ "€"_-;_-@_-'
FMT_INT='#,##0'; FMT_PCT='0.0%'

def mf(c): return PatternFill('solid',fgColor=c)
def mn(c=BLK,b=False,sz=11): return Font(name='Calibri',size=sz,bold=b,color=c)
def ma(h='center'): return Alignment(horizontal=h,vertical='center')

ST={
    'hB':(mn(WHT,True),mf(BLK),ma()), 'hO':(mn(WHT,True),mf(ORG),ma()),
    'tx':(mn(),None,ma()),             'eu':(mn(),None,ma()),
    'in':(mn(),None,ma()),             'aT':(mn(),mf(GRY),ma()),
    'aE':(mn(),mf(GRY),ma()),         'aI':(mn(),mf(GRY),ma()),
    'dG':(mn(),mf(GRN),ma()),         'dR':(mn(),mf(RED),ma()),
    'rT':(mn(BLK,True,13),None,None), 'rH':(mn(WHT,True),mf(BLK),ma()),
    'rO':(mn(WHT,True),mf(ORG),ma()),
}
r2=lambda n:round(float(n),2) if n is not None else None

COLS=['tienda','COMPARABLE','temporada','Vender_en','seccion','gama',
      'articulo','codart','color','neto','qty','fecha']

def mvend(v):
    if v=='SS26': return'SS26'
    if v=='NOS CONTINUATIVO': return'NOS'
    if v=='FIN EXISTENCIAS': return'FE'
    if v=='CEREMONIA': return'CEREMONIA'
    return None

# ── Procesamiento (se ejecuta en hilo) ────────────────────────────────────────
def procesar(job_id, file_bytes, tipo, semana, anio):
    def upd(p, msg):
        with JOBS_LOCK:
            JOBS[job_id]['progress'] = p
            JOBS[job_id]['message']  = msg

    try:
        upd(5, 'Leyendo archivo...')

        # ── Lectura streaming ────────────────────────────────────────────────
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_name = wb.sheetnames[0]
        for sn in wb.sheetnames:
            if any(x in sn.lower() for x in ['maestro','full','outlet']):
                sheet_name = sn; break
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h else '' for h in next(rows_iter)]

        def fi(name):
            for i,h in enumerate(header):
                if h.lower() == name.lower(): return i
            return -1

        I = {c: fi(c) for c in COLS}
        missing = [c for c in COLS if I[c] < 0]
        if missing:
            raise ValueError(f'Columnas no encontradas: {missing}')

        # Acumuladores
        mFT=defaultdict(lambda:[0.,0.,0.,0.])
        mFS=defaultdict(lambda:[0.,0.,0.,0.])
        mGC=defaultdict(lambda:[0.,0.,0.,0.])
        mGT=defaultdict(lambda:[0.,0.,0.,0.])
        mGTd=defaultdict(lambda:[0.,0.,0.,0.])
        mRS=defaultdict(lambda:[0.,0.])
        mRST=defaultdict(float)
        mRC=defaultdict(lambda:[0.,0.])
        mRCT=defaultdict(float)
        mTC={}; mTV={}; mVT={}; mTTc={}

        def acc(d,k,n,q,f):
            e=d[k]
            if f==2026: e[0]+=n;e[1]+=q
            else:       e[2]+=n;e[3]+=q

        def acT(d,k,n,q,dims):
            if k not in d: d[k]=[0.,0.]+list(dims)
            d[k][0]+=n; d[k][1]+=q

        upd(10, 'Procesando datos...')
        count = 0
        for row in rows_iter:
            try:
                fecha=int(row[I['fecha']] or 0)
                if fecha not in (2025,2026): continue
                td  =str(row[I['tienda']]     or'').strip()
                comp=str(row[I['COMPARABLE']] or'').strip()
                if comp in('0','None','nan'): comp=''
                tmp =str(row[I['temporada']]  or'').strip()
                vnd =str(row[I['Vender_en']]  or'').strip()
                sec =str(row[I['seccion']]    or'').strip()
                gam =str(row[I['gama']]       or'').strip()
                art =str(row[I['articulo']]   or'').strip()
                cod =str(row[I['codart']]      or'').strip()
                col =str(row[I['color']]      or'').strip()
                neto=float(row[I['neto']]     or 0)
                qty =float(row[I['qty']]      or 0)
                isC =comp in('SI','SI - Reformada')

                acc(mFT, td,                 neto,qty,fecha)
                acc(mFS, f'{td}|{sec}',      neto,qty,fecha)
                if isC: acc(mGC,f'{sec}|{gam}',neto,qty,fecha)
                acc(mGT, f'{sec}|{gam}',     neto,qty,fecha)
                acc(mGTd,f'{td}|{sec}|{gam}',neto,qty,fecha)

                r=mRS[sec]
                if fecha==2026: r[0]+=neto
                else:           r[1]+=neto
                if fecha==2026:
                    tg=mvend(vnd)
                    if tg: mRST[f'{sec}|{tg}']+=neto
                if isC:
                    r=mRC[sec]
                    if fecha==2026: r[0]+=neto
                    else:           r[1]+=neto
                    if fecha==2026:
                        tg=mvend(vnd)
                        if tg: mRCT[f'{sec}|{tg}']+=neto

                acT(mTC, f'{fecha}|{sec}|{tmp}|{vnd}|{gam}|{cod}|{art}',
                    neto,qty,[fecha,sec,tmp,vnd,gam,cod,art])
                acT(mTV, f'{fecha}|{sec}|{tmp}|{vnd}|{gam}|{cod}|{art}|{col}',
                    neto,qty,[fecha,sec,tmp,vnd,gam,cod,art,col])
                acT(mVT, f'{td}|{fecha}|{sec}|{tmp}|{vnd}|{gam}|{cod}|{art}|{col}',
                    neto,qty,[td,fecha,sec,tmp,vnd,gam,cod,art,col])
                acT(mTTc,f'{td}|{fecha}|{sec}|{tmp}|{vnd}|{gam}|{cod}|{art}',
                    neto,qty,[td,fecha,sec,tmp,vnd,gam,cod,art])
                count+=1
            except: pass

            if count % 20000 == 0:
                upd(10 + min(40, count//3000), f'Procesando... {count:,} filas')

        wb.close()
        del file_bytes; gc.collect()
        upd(55, 'Calculando pestañas...')

        M=dict(mFT=mFT,mFS=mFS,mGC=mGC,mGT=mGT,mGTd=mGTd,
               mRS=mRS,mRST=mRST,mRC=mRC,mRCT=mRCT,
               mTC=mTC,mTV=mTV,mVT=mVT,mTTc=mTTc)

        # ── Escritura ────────────────────────────────────────────────────────
        upd(60, 'Generando Excel...')
        buf, fname = generar_excel(M, tipo, semana, anio)
        del M; gc.collect()

        import base64
        b64 = base64.b64encode(buf.read()).decode('ascii')
        del buf; gc.collect()
        upd(100, 'Listo')
        with JOBS_LOCK:
            JOBS[job_id]['status'] = 'done'
            JOBS[job_id]['b64']    = b64
            JOBS[job_id]['fname']  = fname

    except Exception as e:
        tb = traceback.format_exc()
        with JOBS_LOCK:
            JOBS[job_id]['status']  = 'error'
            JOBS[job_id]['error']   = str(e)
            JOBS[job_id]['message'] = str(e)
        print(f"ERROR job {job_id}:\n{tb}")

# ── AOAs ──────────────────────────────────────────────────────────────────────
def aoa_resumen(mRS,mRST,mRC,mRCT):
    TC=['SS26','NOS','FE','CEREMONIA']
    def block(mS,mST):
        tot26=sum(e[0] for e in mS.values())
        rows=[]
        for sec in sorted(mS,key=lambda s:-mS[s][0]):
            e=mS[sec];n26=r2(e[0]);n25=r2(e[1])
            pct=round(n26/tot26,4) if tot26 else 0
            ts=[r2(mST.get(f'{sec}|{tc}',0)) for tc in TC]
            tp=[round(t/n26,4) if n26 else 0 for t in ts]
            rows.append([sec,n26,n25,r2(n26-n25),pct,None,sec]+ts+[r2(sum(ts)),None,sec]+tp)
        tn26=r2(sum(e[0] for e in mS.values()))
        tn25=r2(sum(e[1] for e in mS.values()))
        tt=[r2(sum(mST.get(f'{s}|{tc}',0) for s in mS)) for tc in TC]
        ttp=[round(t/tn26,4) if tn26 else 0 for t in tt]
        rows.append(['Total general',tn26,tn25,r2(tn26-tn25),1.0,
                     None,'Total general']+tt+[r2(sum(tt)),None,'Total general']+ttp)
        return rows
    hdr=['SECCIÓN',2026,2025,'DIF NETO','%',None,'SECCIÓN','SS26','NOS','FE','CEREMONIA','Total',
         None,'SECCIÓN','SS26%','NOS%','FE%','CER%']
    return [['TOTALES'],[],[],hdr]+block(mRS,mRST)+\
           [[],['COMPARABLES'],[],[],hdr]+block(mRC,mRCT)

def aoa_fac_tienda(m):
    rows=[['tienda',2026,2025,'dif neto']]
    for td,e in sorted(m.items(),key=lambda x:-x[1][0]):
        n26=r2(e[0]);n25=r2(e[2]);rows.append([td,n26,n25,r2(n26-n25)])
    return rows

def aoa_fac_sec(m):
    rows=[['tienda','seccion',2026,2025,'DIF NETO']]
    for k,e in sorted(m.items(),key=lambda x:-x[1][0]):
        td,sec=k.split('|',1);n26=r2(e[0]);n25=r2(e[2])
        rows.append([td,sec,n26,n25,r2(n26-n25)])
    return rows

def aoa_gama_comp(m):
    rows=[['seccion','gama','NETO26','NETO25','DIF NETO','QTY26','QTY25','DIF QTY','PM 26','PM 25','DIF PM']]
    for k,e in sorted(m.items(),key=lambda x:-x[1][0]):
        sec,gam=k.split('|',1)
        n26=r2(e[0]);n25=r2(e[2]);q26=round(e[1]);q25=round(e[3])
        pm26=r2(n26/q26) if q26 else None
        pm25=r2(n25/q25) if q25 else None
        dpm=r2(pm26-pm25) if pm26 is not None and pm25 is not None else None
        rows.append([sec,gam,n26,n25,r2(n26-n25),q26,q25,q26-q25,pm26,pm25,dpm])
    return rows

def aoa_top_cod(m):
    rows=[['fecha','seccion','temporada','Vender_en','gama','codart','articulo','Suma de neto','Suma de qty']]
    for e in sorted(m.values(),key=lambda x:-x[0]):
        rows.append([e[2],e[3],e[4],e[5],e[6],e[7],e[8],r2(e[0]),round(e[1])])
    return rows

def aoa_gama_tot(m):
    rows=[['seccion','gama','NETO26','NETO25','DIF NETO','QTY26','QTY25','DIF QTY']]
    for k,e in sorted(m.items(),key=lambda x:-x[1][0]):
        sec,gam=k.split('|',1)
        n26=r2(e[0]);n25=r2(e[2]);q26=round(e[1]);q25=round(e[3])
        rows.append([sec,gam,n26,n25,r2(n26-n25),q26,q25,q26-q25])
    return rows

def aoa_gama_tienda(m):
    rows=[['tienda','seccion','gama','NETO 26','NETO 25','DIF NETO','QTY 26','QTY 25','DIF QTY']]
    for k,e in sorted(m.items(),key=lambda x:-x[1][0]):
        parts=k.split('|',2)
        td=parts[0];sec=parts[1];gam=parts[2] if len(parts)>2 else ''
        n26=r2(e[0]);n25=r2(e[2]);q26=round(e[1]);q25=round(e[3])
        rows.append([td,sec,gam,n26,n25,r2(n26-n25),q26,q25,q26-q25])
    return rows

def aoa_top_vent_cia(m):
    rows=[['fecha','seccion','temporada','Vender_en','gama','codart','articulo','color','Suma de neto','Suma de qty']]
    for e in sorted(m.values(),key=lambda x:-x[0]):
        rows.append([e[2],e[3],e[4],e[5],e[6],e[7],e[8],e[9],r2(e[0]),round(e[1])])
    return rows

def aoa_top_venta_tienda(m):
    rows=[['tienda','fecha','seccion','temporada','Vender_en','gama','codart','articulo','color','Suma de neto','Suma de qty']]
    for e in sorted(m.values(),key=lambda x:-x[0]):
        rows.append([e[2],e[3],e[4],e[5],e[6],e[7],e[8],e[9],e[10],r2(e[0]),round(e[1])])
    return rows

def aoa_top_tienda_cod(m):
    rows=[['tienda','fecha','seccion','temporada','Vender_en','gama','codart','articulo','Suma de neto','Suma de qty']]
    for e in sorted(m.values(),key=lambda x:-x[0]):
        rows.append([e[2],e[3],e[4],e[5],e[6],e[7],e[8],e[9],r2(e[0]),round(e[1])])
    return rows

# ── Escritura write_only ──────────────────────────────────────────────────────
SHEET_DEFS={
    'FACTURACIÓN TIENDAS':   [('tienda',47,None,False),(2026,13,FMT_EUR,False),(2025,13,FMT_EUR,False),('dif neto',13,FMT_EUR,True)],
    'FACTURACIÓN SECCIÓN':   [('tienda',47,None,False),('seccion',16,None,False),(2026,13,FMT_EUR,False),(2025,13,FMT_EUR,False),('DIF NETO',13,FMT_EUR,True)],
    'GAMAS CIA COMPARABLES': [('seccion',14,None,False),('gama',17,None,False),('NETO26',13,FMT_EUR,False),('NETO25',13,FMT_EUR,False),('DIF NETO',12,FMT_EUR,True),('QTY26',10,FMT_INT,False),('QTY25',11,FMT_INT,False),('DIF QTY',11,FMT_INT,True),('PM 26',11,FMT_EUR,False),('PM 25',11,FMT_EUR,False),('DIF PM',11,FMT_EUR,True)],
    'TOP CIA COD':           [('fecha',10,None,False),('seccion',14,None,False),('temporada',17,None,False),('Vender_en',17,None,False),('gama',15,None,False),('codart',12,None,False),('articulo',57,None,False),('Suma de neto',13,FMT_EUR,False),('Suma de qty',11,FMT_INT,False)],
    'GAMAS CIA TOTALES':     [('seccion',14,None,False),('gama',17,None,False),('NETO26',13,FMT_EUR,False),('NETO25',13,FMT_EUR,False),('DIF NETO',12,FMT_EUR,True),('QTY26',10,FMT_INT,False),('QTY25',11,FMT_INT,False),('DIF QTY',11,FMT_INT,True)],
    'GAMAS TIENDA':          [('tienda',47,None,False),('seccion',14,None,False),('gama',17,None,False),('NETO 26',12,FMT_EUR,False),('NETO 25',12,FMT_EUR,False),('DIF NETO',12,FMT_EUR,True),('QTY 26',10,FMT_INT,False),('QTY 25',11,FMT_INT,False),('DIF QTY',11,FMT_INT,True)],
    'TOP VENTAS CIA':        [('fecha',10,None,False),('seccion',14,None,False),('temporada',17,None,False),('Vender_en',17,None,False),('gama',15,None,False),('codart',12,None,False),('articulo',57,None,False),('color',22,None,False),('Suma de neto',13,FMT_EUR,False),('Suma de qty',11,FMT_INT,False)],
    'TOP VENTA TIENDA':      [('tienda',47,None,False),('fecha',10,None,False),('seccion',14,None,False),('temporada',17,None,False),('Vender_en',17,None,False),('gama',15,None,False),('codart',12,None,False),('articulo',57,None,False),('color',22,None,False),('Suma de neto',13,FMT_EUR,False),('Suma de qty',11,FMT_INT,False)],
    'TOP TIENDA COD':        [('tienda',47,None,False),('fecha',10,None,False),('seccion',14,None,False),('temporada',17,None,False),('Vender_en',17,None,False),('gama',15,None,False),('codart',12,None,False),('articulo',57,None,False),('Suma de neto',13,FMT_EUR,False),('Suma de qty',11,FMT_INT,False)],
}

def wc(ws,val,sk,fmt=None):
    fnt,fil,aln=ST[sk]
    c=WriteOnlyCell(ws,value=val)
    c.font=fnt
    if fil: c.fill=fil
    if aln: c.alignment=aln
    if fmt and val is not None: c.number_format=fmt
    return c

def write_data_sheet(wb,name,aoa):
    ws=wb.create_sheet(name)
    col_defs=SHEET_DEFS[name]
    for ci,(_,w,_,_) in enumerate(col_defs,1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.append([wc(ws,lbl,'hO' if orange else 'hB') for lbl,_,_,orange in col_defs])
    for ri,row in enumerate(aoa[1:],2):
        alt=ri%2==0; out=[]
        for ci,(cn,_,fmt,_) in enumerate(col_defs):
            val=row[ci] if ci<len(row) else None
            is_dif='DIF' in str(cn).upper()
            if is_dif and isinstance(val,(int,float)) and val is not None:
                sk='dG' if val>0 else('dR' if val<0 else('aE' if alt else 'eu'))
            elif fmt==FMT_EUR: sk='aE' if alt else 'eu'
            elif fmt==FMT_INT: sk='aI' if alt else 'in'
            else:              sk='aT' if alt else 'tx'
            out.append(wc(ws,val,sk,fmt))
        ws.append(out)

def write_resumen(wb,aoa):
    ws=wb.create_sheet('RESUMEN')
    for col,w in {'A':16,'B':14,'C':14,'D':12,'E':11,'G':16,'H':14,'I':13,'J':12,'K':12,'L':14,'N':13,'O':7,'P':7,'Q':7,'R':12}.items():
        ws.column_dimensions[col].width=w
    for row in aoa:
        if not row: ws.append([None]); continue
        if row[0] in('TOTALES','COMPARABLES') and all(v is None for v in row[1:]):
            c=WriteOnlyCell(ws,value=row[0]); c.font=ST['rT'][0]; ws.append([c]); continue
        is_hdr=row[0]=='SECCIÓN'; is_tot=row[0]=='Total general'; out=[]
        for ci,val in enumerate(row,1):
            if val is None: out.append(None); continue
            if is_hdr:   c=wc(ws,val,'rH')
            elif is_tot:
                c=wc(ws,val,'rO')
                if ci in(2,3,4,8,9,10,11,12): c.number_format=FMT_EUR
                if ci==5 or ci in(15,16,17,18): c.number_format=FMT_PCT
            else:
                c=wc(ws,val,'tx')
                if ci in(2,3,4,8,9,10,11,12): c.number_format=FMT_EUR
                if ci==5 or ci in(15,16,17,18): c.number_format=FMT_PCT
                if ci==4 and isinstance(val,(int,float)):
                    c.fill=mf(GRN if val>=0 else RED)
            out.append(c)
        ws.append(out)

def generar_excel(M,tipo,semana,anio):
    wb=Workbook(write_only=True)
    write_resumen(wb,aoa_resumen(M['mRS'],M['mRST'],M['mRC'],M['mRCT']))
    for name,fn,key in [
        ('FACTURACIÓN TIENDAS',   aoa_fac_tienda,      'mFT'),
        ('FACTURACIÓN SECCIÓN',   aoa_fac_sec,         'mFS'),
        ('GAMAS CIA COMPARABLES', aoa_gama_comp,       'mGC'),
        ('TOP CIA COD',           aoa_top_cod,         'mTC'),
        ('GAMAS CIA TOTALES',     aoa_gama_tot,        'mGT'),
        ('GAMAS TIENDA',          aoa_gama_tienda,     'mGTd'),
        ('TOP VENTAS CIA',        aoa_top_vent_cia,    'mTV'),
        ('TOP VENTA TIENDA',      aoa_top_venta_tienda,'mVT'),
        ('TOP TIENDA COD',        aoa_top_tienda_cod,  'mTTc'),
    ]:
        aoa=fn(M[key]); write_data_sheet(wb,name,aoa); del aoa; gc.collect()
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    del wb; gc.collect()
    return buf, f'PESOS_{tipo}_W{semana:02d}_{anio}.xlsx'

# ── Rutas ─────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    now=datetime.now()
    return render_template_string(HTML,sem=now.isocalendar()[1],anio=now.year)

@app.route('/iniciar', methods=['POST'])
def iniciar():
    """Recibe fichero como JSON base64, lanza hilo, devuelve job_id."""
    cleanup_old_jobs()
    import base64
    data     = request.get_json(force=True)
    tipo     = data.get('tipo','FULL_PRICE')
    sem      = int(data.get('semana',1))
    anio     = int(data.get('anio',datetime.now().year))
    b64      = data.get('file_b64','')
    if not b64: return jsonify(error='No se recibió archivo'), 400
    file_bytes = base64.b64decode(b64)
    job_id = str(uuid.uuid4())

    with JOBS_LOCK:
        JOBS[job_id] = {
            'status':'running', 'progress':0,
            'message':'Iniciando...', 'b64':None,
            'fname':None, 'error':None, 'ts':time.time()
        }

    t = threading.Thread(target=procesar, args=(job_id,file_bytes,tipo,sem,anio), daemon=True)
    t.start()
    return jsonify(job_id=job_id)

@app.route('/estado/<job_id>')
def estado(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job: return jsonify(error='Job no encontrado'),404
    resp = dict(
        status   = job['status'],
        progress = job['progress'],
        message  = job['message'],
        fname    = job['fname'],
        error    = job['error']
    )
    if job['status'] == 'done':
        resp['b64'] = job.get('b64','')
    return jsonify(resp)

@app.route('/descargar/<job_id>')
def descargar(job_id):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job or job['status']!='done':
        return 'No disponible', 404
    buf   = job['result_buf']
    fname = job['fname']
    buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)

# ── HTML ──────────────────────────────────────────────────────────────────────
HTML='''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Generador PESOS</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{--ink:#0f0f14;--paper:#f5f4ef;--accent:#c8f035;--mid:#7a7a8a;--border:#d8d7cf;--card:#fff;--orange:#e97132;--green:#c6efce;--green-dk:#276221;--red:#ffc7ce}
body{font-family:'DM Sans',sans-serif;background:var(--paper);color:var(--ink);min-height:100vh}
header{background:var(--ink);padding:18px 36px;display:flex;align-items:center;justify-content:space-between}
.logo{font-family:'DM Mono',monospace;font-size:12px;font-weight:500;color:var(--accent);letter-spacing:.14em;text-transform:uppercase}
.pill{font-family:'DM Mono',monospace;font-size:11px;color:#555;background:#1a1a22;padding:5px 12px;border-radius:20px}
main{max-width:760px;margin:0 auto;padding:36px 20px 80px}
.week-row{display:flex;gap:20px;align-items:center;margin-bottom:28px;padding:18px 22px;background:var(--card);border-radius:12px;border:1px solid var(--border)}
.wlbl{display:flex;flex-direction:column;gap:4px;align-items:center}
.wlbl label{font-size:10px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--mid)}
.wlbl input{font-family:'DM Mono',monospace;font-size:22px;font-weight:500;width:68px;border:none;border-bottom:2px solid var(--ink);background:transparent;color:var(--ink);padding:2px 0;text-align:center;outline:none;-moz-appearance:textfield}
.wlbl input::-webkit-inner-spin-button{-webkit-appearance:none}
.wlbl input:focus{border-color:var(--accent)}
.wsep{font-family:'DM Mono',monospace;font-size:20px;color:var(--border)}
.drop-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:28px}
.dz{border:1.5px dashed var(--border);border-radius:12px;padding:26px 18px;text-align:center;cursor:pointer;transition:all .2s;background:var(--card);position:relative;min-height:145px;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:10px}
.dz:hover,.dz.over{border-color:var(--ink);background:#fafaf6}
.dz.loaded{border-color:var(--accent);border-style:solid;background:#f6ffe8}
.dz input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.dz-icon{font-size:26px}.dz-lbl{font-size:10px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:var(--mid)}
.dz-sub{font-size:11px;color:#bbb}.dz-opt{font-size:10px;color:#ccc;font-style:italic}
.dz-name{font-family:'DM Mono',monospace;font-size:10px;background:var(--accent);color:var(--ink);padding:2px 8px;border-radius:3px;font-weight:600;max-width:180px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.dz-meta{font-family:'DM Mono',monospace;font-size:9px;color:var(--mid);line-height:1.8}
.dz-clr{position:absolute;top:8px;right:10px;background:none;border:none;font-size:14px;color:#bbb;cursor:pointer;padding:3px 6px;border-radius:4px;z-index:2}
.dz-clr:hover{color:var(--ink);background:#f0ede4}
.fmt-bar{display:flex;gap:7px;flex-wrap:wrap;margin-bottom:24px;padding:10px 14px;background:var(--card);border-radius:10px;border:1px solid var(--border);align-items:center}
.fmt-bar>span:first-child{font-size:10px;font-weight:600;color:var(--mid)}
.chip{font-size:10px;font-weight:600;padding:3px 8px;border-radius:4px}
.c-blk{background:var(--ink);color:var(--accent)}.c-org{background:var(--orange);color:#fff}
.c-grn{background:var(--green);color:var(--green-dk)}.c-red{background:var(--red);color:#9c0006}
#prog{display:none;padding:20px 22px;background:var(--card);border-radius:12px;border:1px solid var(--border);margin-bottom:24px}
.prog-top{display:flex;justify-content:space-between;margin-bottom:8px}
.prog-lbl{font-family:'DM Mono',monospace;font-size:11px;color:var(--mid)}
.prog-pct{font-family:'DM Mono',monospace;font-size:12px;font-weight:500}
.prog-track{height:4px;background:var(--border);border-radius:2px;overflow:hidden}
.prog-fill{height:100%;background:var(--ink);border-radius:2px;transition:width .5s ease;width:0%}
#err{display:none;background:#fff5f5;border:1px solid #ffd0d0;border-radius:10px;padding:14px 18px;font-size:12px;color:#cc2222;margin-bottom:20px;font-family:'DM Mono',monospace;white-space:pre-wrap}
#btn{width:100%;padding:16px;background:var(--ink);color:var(--accent);border:none;border-radius:12px;font-family:'DM Mono',monospace;font-size:13px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;cursor:pointer;transition:all .2s;display:flex;align-items:center;justify-content:center;gap:10px}
#btn:hover:not(:disabled){background:#1e1e2e;transform:translateY(-1px)}
#btn:disabled{opacity:.3;cursor:not-allowed;transform:none}
.results{display:flex;flex-direction:column;gap:12px;margin-top:20px}
.rc{background:var(--ink);border-radius:14px;padding:22px 26px;color:#fff;display:flex;align-items:center;justify-content:space-between;gap:20px}
.ri{display:flex;align-items:center;gap:14px}
.ricon{font-size:26px}.rname{font-family:'DM Mono',monospace;font-size:13px;color:var(--accent);font-weight:500}
.rsub{font-size:11px;color:#666;margin-top:2px}
.dl{padding:10px 20px;background:var(--accent);color:var(--ink);border:none;border-radius:8px;font-family:'DM Mono',monospace;font-size:12px;font-weight:700;text-transform:uppercase;cursor:pointer;white-space:nowrap;text-decoration:none;display:inline-block}
.dl:hover{background:#d4ff40}
</style>
</head>
<body>
<header>
  <span class="logo">⬡ Generador PESOS</span>
  <span class="pill" id="wpill">W__ · ____</span>
</header>
<main>
  <div class="week-row">
    <div class="wlbl"><label>Semana</label><input type="number" id="isem" min="1" max="53" value="{{ sem }}"/></div>
    <span class="wsep">/</span>
    <div class="wlbl"><label>Año</label><input type="number" id="ianio" min="2020" max="2099" value="{{ anio }}"/></div>
  </div>
  <div class="drop-grid">
    <div class="dz" id="dz-fp"><input type="file" id="f-fp" accept=".xlsx,.xls"/>
      <div class="dz-icon">📋</div><div class="dz-lbl">Maestro Full Price</div><div class="dz-sub">Arrastra aquí o haz clic</div></div>
    <div class="dz" id="dz-out"><input type="file" id="f-out" accept=".xlsx,.xls"/>
      <div class="dz-icon">🏪</div><div class="dz-lbl">Maestro Outlet</div><div class="dz-sub">Arrastra aquí o haz clic</div><div class="dz-opt">opcional</div></div>
  </div>
  <div class="fmt-bar">
    <span>FORMATO:</span>
    <span class="chip c-blk">Cabeceras negras</span><span class="chip c-org">DIF naranja</span>
    <span class="chip c-grn">▲ Positivo</span><span class="chip c-red">▼ Negativo</span>
  </div>
  <div id="err"></div>
  <div id="prog">
    <div class="prog-top">
      <span class="prog-lbl" id="plbl">Procesando...</span>
      <span class="prog-pct" id="ppct">0%</span>
    </div>
    <div class="prog-track"><div class="prog-fill" id="pfill"></div></div>
  </div>
  <button id="btn" disabled>→ &nbsp;Generar PESOS</button>
  <div class="results" id="results"></div>
</main>
<script>
const files={};const $=id=>document.getElementById(id);
function upPill(){$('wpill').textContent=`W${String($('isem').value).padStart(2,'0')} · ${$('ianio').value}`;}
function chkReady(){$('btn').disabled=!Object.keys(files).length;}
function showErr(m){const e=$('err');e.textContent=m;e.style.display='block';}
function hideErr(){$('err').style.display='none';}
function setProg(p,l){$('pfill').style.width=p+'%';$('ppct').textContent=p+'%';if(l)$('plbl').textContent=l;}

function setupDz(key){
  const dz=$(`dz-${key}`);
  dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('over');});
  dz.addEventListener('dragleave',()=>dz.classList.remove('over'));
  dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('over');if(e.dataTransfer.files[0])setFile(e.dataTransfer.files[0],key);});
  $(`f-${key}`).addEventListener('change',e=>{if(e.target.files[0])setFile(e.target.files[0],key);});
}
function setFile(file,key){
  files[key]=file;const dz=$(`dz-${key}`);dz.classList.add('loaded');
  dz.innerHTML=`<button class="dz-clr" onclick="clrFile('${key}')">✕</button>
    <input type="file" id="f-${key}" accept=".xlsx,.xls"/>
    <div class="dz-icon">${key==='fp'?'📋':'🏪'}</div>
    <div class="dz-lbl">${key==='fp'?'Full Price':'Outlet'}</div>
    <div class="dz-name">${file.name}</div>
    <div class="dz-meta">${(file.size/1024/1024).toFixed(1)} MB</div>`;
  $(`f-${key}`).addEventListener('change',e=>{if(e.target.files[0])setFile(e.target.files[0],key);});
  hideErr();chkReady();
}
function clrFile(key){
  delete files[key];const dz=$(`dz-${key}`);dz.classList.remove('loaded');
  dz.innerHTML=`<input type="file" id="f-${key}" accept=".xlsx,.xls"/>
    <div class="dz-icon">${key==='fp'?'📋':'🏪'}</div>
    <div class="dz-lbl">${key==='fp'?'Maestro Full Price':'Maestro Outlet'}</div>
    <div class="dz-sub">Arrastra aquí o haz clic</div>
    ${key==='out'?'<div class="dz-opt">opcional</div>':''}`;
  $(`f-${key}`).addEventListener('change',e=>{if(e.target.files[0])setFile(e.target.files[0],key);});
  chkReady();
}
setupDz('fp');setupDz('out');
$('isem').addEventListener('input',upPill);$('ianio').addEventListener('input',upPill);upPill();

async function waitForJob(jobId,icon,tipo,sem,anio){
  return new Promise((resolve,reject)=>{
    const poll=setInterval(async()=>{
      try{
        const r=await fetch(`/estado/${jobId}`);
        const j=await r.json();
        setProg(j.progress, j.message);
        if(j.status==='done'){
          clearInterval(poll);
          resolve({jobId,icon,tipo,fname:j.fname,b64:j.b64,sem,anio});
        } else if(j.status==='error'){
          clearInterval(poll);
          reject(new Error(j.error||'Error desconocido'));
        }
      }catch(e){clearInterval(poll);reject(e);}
    },2000);
  });
}

$('btn').addEventListener('click',async()=>{
  hideErr();$('results').innerHTML='';$('btn').disabled=true;
  $('prog').style.display='block';setProg(0,'Subiendo archivo...');

  const sem=$('isem').value,anio=$('ianio').value;
  const promises=[];

  try{
    for(const[key,file]of Object.entries(files)){
      const tipo=key==='fp'?'FULL_PRICE':'OUTLET';
      const icon=key==='fp'?'📋':'🏪';
      // Convertir fichero a base64 y enviar como JSON
      const b64=await new Promise((res,rej)=>{
        const rd=new FileReader();
        rd.onload=()=>res(rd.result.split(',')[1]);
        rd.onerror=rej;
        rd.readAsDataURL(file);
      });
      const r=await fetch('/iniciar',{
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body:JSON.stringify({file_b64:b64,tipo,semana:sem,anio})
      });
      if(!r.ok)throw new Error(await r.text());
      const {job_id}=await r.json();
      promises.push(waitForJob(job_id,icon,tipo,sem,anio));
    }

    // Esperar todos los trabajos
    const results=await Promise.all(promises);

    setProg(100,'¡Listo!');
    setTimeout(()=>$('prog').style.display='none',400);

    $('results').innerHTML=results.map(({jobId,icon,tipo,fname,b64})=>`
      <div class="rc">
        <div class="ri"><span class="ricon">${icon}</span>
          <div><div class="rname">${fname}</div>
          <div class="rsub">${tipo==='FULL_PRICE'?'Full Price':'Outlet'} · 10 pestañas</div></div></div>
        <a class="dl" id="dl-${jobId}" download="${fname}">↓ Descargar</a>
      </div>`).join('');
  // Crear blob URLs desde base64 y asignar a los botones
  results.forEach(({jobId,fname,b64})=>{
    const bytes=atob(b64);
    const arr=new Uint8Array(bytes.length);
    for(let i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
    const blob=new Blob([arr],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
    const el=document.getElementById('dl-'+jobId);
    if(el) el.href=URL.createObjectURL(blob);
  });
  }catch(err){
    $('prog').style.display='none';
    showErr('Error: '+err.message);
  }
  $('btn').disabled=false;
});
</script>
</body>
</html>'''

if __name__=='__main__':
    port=int(os.environ.get('PORT',5000))
    app.run(host='0.0.0.0',port=port,debug=False)
